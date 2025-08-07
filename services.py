import httpx
import openpyxl

from typing import IO


async def get_id_nickname(name):
    name = str(name)
    name = name.strip()
    if len(name.split()) > 1:
        return f"Пользователь с ником {name} не найден"
    data = None

    try:
        response = httpx.get(f"https://api.tanki.su/wot/account/list/?application_id=6e76bdefe5b324d811d62bd6bc861e02&search={name}")
        data = response.json()
    except httpx.ReadTimeout:
        response = httpx.get(f"https://api.tanki.su/wot/account/list/?application_id=6e76bdefe5b324d811d62bd6bc861e02&search={name}")
        data = response.json()


    if data["status"] != "ok":
        return f"Пользователь с ником {name} не найден"

    if not data["data"]:
        return f"Пользователь с ником {name} не найден"

    nickname = data["data"][0]["nickname"]
    if nickname.lower() != name.lower():
        return f"Пользователь с ником {name} не найден"

    id_ = data["data"][0]["account_id"]
    return str(id_)


async def excel_parser(file: IO[bytes]):
    wb = openpyxl.load_workbook(file, keep_vba=True)
    sheet = wb.active
    max_rows = sheet.max_row

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    if max_rows > 500:
        raise TimeoutError

    for i in range(1, max_rows + 1):
        nickname = sheet.cell(row=i, column=1)
        if not nickname.value:
            continue
        id_ = await get_id_nickname(nickname.value)
        new_sheet.cell(row=i, column=1).value = id_
        new_sheet.cell(row=i, column=2).value = sheet.cell(row=i, column=2).value

    new_wb.save("test.xlsx")
    new_wb.close()
    wb.close()


async def txt_parser(file):
    id_list = []

    with open(file, "r") as f:
        for line in f:
            id_ = await get_id_nickname(line)
            id_list.append(id_ + "\n")

    with open("data_list_id.txt", "a") as f:
        f.writelines(id_list)

